# -*- coding: utf-8 -*-
# @Time    : 2020/7/4
# @Author  : Lart Pang
# @Email   : lartpang@163.com
# @File    : recoder.py
# @Project : utils/recoder.py
# @GitHub  : https://github.com/lartpang
import functools
import os
from datetime import datetime

from openpyxl import Workbook, load_workbook
from torch.utils.tensorboard import SummaryWriter
from torchvision.utils import make_grid

from utils.misc import check_mkdir, construct_print


class TBRecorder(object):
    def __init__(self, tb_path):
        check_mkdir(dir_path=tb_path)

        self.tb = SummaryWriter(tb_path)

    def record_curve(self, name, data, curr_iter):
        if not isinstance(data, (tuple, list)):
            self.tb.add_scalar(f"data/{name}", data, curr_iter)
        else:
            for idx, data_item in enumerate(data):
                self.tb.add_scalar(f"data/{name}_{idx}", data_item[name], curr_iter)

    def record_image(self, name, data, curr_iter):
        data_grid = make_grid(data, nrow=data.size(0), padding=5)
        self.tb.add_image(name, data_grid, curr_iter)

    def close_tb(self):
        self.tb.close()


class XLSXRecoder(object):
    def __init__(self, xlsx_path):
        self.dataset_list = ["DUTS", "DUT-OMRON", "HKU-IS", "ECSSD", "PASCAL-S", "SOC"]
        self.dataset_num_list = [5019, 5168, 1447, 1000, 850, 1200]
        self.metric_list = ["MAXF", "MEANF", "MAE"]

        self.path = xlsx_path
        if not os.path.exists(self.path):
            self.create_xlsx()

    def create_xlsx(self):
        num_metrics = len(self.metric_list)
        num_datasets = len(self.dataset_list)

        # 创建一个Workbook对象
        wb = Workbook()
        # 创建一个Sheet对象
        sheet = wb.create_sheet(title="Results", index=0)
        # 获取活动的sheet
        sheet["A1"] = "name_dataset"
        sheet["A2"] = "num_dataset"

        for i, dataset_name in enumerate(self.dataset_list):
            if (i * num_metrics + 1) // 26 == 0:
                start_region_idx = f"{chr(ord('A') + (i * num_metrics + 1) % 26)}1"
            else:
                start_region_idx = (
                    f"{chr(ord('A') + (i * num_metrics + 1) // 26 - 1)}"
                    f"{chr(ord('A') + (i * num_metrics + 1) % 26)}1"
                )
            if ((i + 1) * num_metrics) // 26 == 0:
                end_region_idx = f"{chr(ord('A') + ((i + 1) * num_metrics) % 26)}1"
            else:
                end_region_idx = (
                    f"{chr(ord('A') + ((i + 1) * num_metrics) // 26 - 1)}"
                    f"{chr(ord('A') + ((i + 1) * num_metrics) % 26)}1"
                )
            region_idx = f"{start_region_idx}:{end_region_idx}"
            sheet.merge_cells(region_idx)  # 合并一行中的几个单元格
            sheet[start_region_idx] = dataset_name.upper()

            # 构造第二行数据
            start_region_idx = start_region_idx.replace("1", "2")
            sheet[start_region_idx] = self.dataset_num_list[i]

        # 构造第三行数据
        third_row = ["metrics"] + self.metric_list * num_datasets
        sheet.append(third_row)

        # 最后保存workbook
        wb.save(self.path)

    def write_xlsx(self, model_name, data):
        """
        向xlsx文件中写入数据

        :param model_name: 模型名字
        :param data: 数据信息，包含数据集名字和对应的测试结果
        """

        num_metrics = len(self.metric_list)
        num_datasets = len(self.dataset_list)

        # 必须先得由前面的部分进行xlsx文件的创建，确保前三行OK满足要求，后面的操作都是从第四行开始的
        wb = load_workbook(self.path)
        assert "Results" in wb.sheetnames, (
            "Please make sure you are " "working with xlsx files " "created by `create_xlsx`"
        )
        sheet = wb["Results"]
        num_cols = num_metrics * num_datasets + 1

        if model_name in sheet["A"]:
            # 说明，该模型已经存在条目中，只需要更新对应的数据集结果即可
            idx_insert_row = sheet["A"].find(model_name)
        else:
            idx_insert_row = len(sheet["A"]) + 1
            sheet.cell(row=idx_insert_row, column=1, value=model_name)

        for dataset_name in data.keys():
            # 遍历每个单元格
            for row in sheet.iter_rows(min_row=1, min_col=2, max_col=num_cols, max_row=1):
                for cell in row:
                    if cell.value == dataset_name.upper():
                        for i in range(num_metrics):
                            matric_name = sheet.cell(row=3, column=cell.column + i).value
                            sheet.cell(
                                row=idx_insert_row,
                                column=cell.column + i,
                                value=data[dataset_name][matric_name.upper()],
                            )
        wb.save(self.path)


def Timer(func):
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        start_time = datetime.now()
        construct_print(f"a new epoch start: {start_time}")
        func(*args, **kwargs)
        construct_print(f"the time of the epoch: {datetime.now() - start_time}")

    return wrapper
